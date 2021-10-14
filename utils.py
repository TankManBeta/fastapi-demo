# -*- coding: utf-8 -*-

"""
    @Author 坦克手贝塔
    @Date 2021/10/10 11:18
"""
import openpyxl
import os
import random


def remove_id(file_path):
    work_book = openpyxl.load_workbook(file_path)
    work_sheet = work_book.active
    for i in range(1, work_sheet.max_row+1):
        work_sheet.cell(i, 1).value = work_sheet.cell(i, 1).value.split('.')[-1]
        work_sheet.cell(i, 1).value = work_sheet.cell(i, 1).value.replace("”“", "” “").replace("” ", "”\n")
    work_book.save(file_path)


def get_sentence(file_path, default=True):
    work_book = openpyxl.load_workbook(file_path)
    work_sheet = work_book.active
    row_num = work_sheet.max_row
    if default:
        rand_int = 1
    else:
        rand_int = random.randint(1, row_num)
    sentence = work_sheet.cell(rand_int, 1).value
    return sentence


def get_img_src(file_path):
    all_images = os.listdir(file_path)
    if "default.jpg" in all_images:
        all_images.remove("default.jpg")
    random_img = random.randint(0, len(all_images) - 1)
    img_src = "http://127.0.0.1:8000/static/images/" + all_images[random_img]
    return img_src


# rename all the picture files
def rename_files(dir_path):
    all_images = os.listdir(dir_path)
    all_images.remove("default.jpg")
    abs_path = os.path.abspath(dir_path)
    for i in range(0, len(all_images)):
        old_filename = abs_path + '\\' + all_images[i]
        new_filename = abs_path + '\\' + str(i) + '.' + all_images[i].split('.')[-1]
        os.rename(old_filename, new_filename)


rename_files('./static/images')
